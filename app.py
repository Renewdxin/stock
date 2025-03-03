import os
import time
from datetime import date, timedelta
from urllib.parse import quote

import pytz
import requests
from apscheduler.schedulers.background import BackgroundScheduler
from flask import Flask, send_from_directory, jsonify, render_template_string, session, redirect, url_for, request
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.exceptions import InvalidFileException
import logging

logging.basicConfig(
    level=logging.INFO,  # 将日志级别从 DEBUG 改为 INFO
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app.log'),  # 添加文件日志
        logging.StreamHandler()  # 保留控制台输出
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 配置 Flask session 密钥和用户登录密码
app.secret_key = os.getenv("FLASK_SECRET_KEY")  # 移除默认值
if not app.secret_key:
    raise ValueError("未设置 FLASK_SECRET_KEY 环境变量")

app.config["USER_PASSWORD"] = os.getenv("USER_PASSWORD")  # 移除默认值
if not app.config["USER_PASSWORD"]:
    raise ValueError("未设置 USER_PASSWORD 环境变量")

app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=10)  # 登录状态有效期 10 分钟

# 全局拦截器：除登录、静态资源之外，所有请求必须登录
@app.before_request
def require_login():
    logger.debug(f"处理请求: {request.endpoint}")
    if request.endpoint is None:
        return
    if request.endpoint in ['login', 'static']:
        return
    if not session.get("logged_in"):
        logger.info("用户未登录，重定向到登录页面")
        return redirect(url_for("login"))

# 登录页面：用户输入密钥后提交验证
@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == app.config["USER_PASSWORD"]:
            session["logged_in"] = True
            session.permanent = True  # 激活 session 的有效期设置
            return redirect(url_for("home"))
        else:
            error = "密钥错误，请重试！"
    login_html = '''
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="utf-8">
        <title>请输入密钥</title>
        <style>
            body {font-family: Arial, sans-serif; text-align: center; margin-top: 100px; background-color: #f4f4f4;}
            .form-container {display: inline-block; padding: 20px; background-color: #fff; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);}
        </style>
    </head>
    <body>
        <div class="form-container">
            <h2>请输入密钥访问网站</h2>
            <form method="post" action="/login">
                <input type="password" name="password" placeholder="请输入密钥" required>
                <button type="submit">提交</button>
            </form>
            <p style="color:red;">{{ error }}</p>
        </div>
    </body>
    </html>
    '''
    return render_template_string(login_html, error=error)

# 可选：注销登录的接口，让用户退出登录
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# 以下内容为原有股票数据更新、文件下载等代码，不变
request_url = "https://query1.finance.yahoo.com/v8/finance/chart/"
request_stock_code = [
    "1797.HK", "0241.HK", "0700.HK", "^HSI", "^IXIC", "9988.HK", "1810.HK", "2318.HK",
    "601318.SS", "002602.SZ", "600276.SS", "000001.SZ", "601127.SS",
    "002594.SZ", "300750.SZ", "002230.SZ", "600436.SS", "000858.SZ",
    "601111.SS", "000001.SS", "399001.SZ", "399006.SZ", "159766.SZ",
    "159875.SZ", "588050.SS", "159928.SZ", "512670.SS", "159901.SZ", "159934.SZ",
    "560080.SS","601985.SS", "161725.SZ", "516110.SS", "512480.SS", "513130.SS",
    "002415.SZ","601899.SS","002085.SZ","513100.SS", "688191.SS"
]
# request_stock_code = ["^HSI", "0241.HK", "0700.HK"]

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) ' \
                  'AppleWebKit/537.36 (KHTML, like Gecko) ' \
                  'Chrome/58.0.3029.110 Safari/537.3'
}

# 配置每个股票的可选列
config_list = {
    "^HSI": ["交易数量"],
    "^IXIC": ["交易数量"],
    "399006.SZ": ["交易数量"],
    "000001.SS": ["交易数量"],
    "399001.SZ": ["交易数量"],
    "0700.HK": ["备注"],
    "1797.HK": ["备注"],
}

def get_stock_data_from_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    stock_data = {}

    for i in range(1, sheet.max_row + 1, 8):
        stock_name = sheet.cell(row=i, column=4).value  # 股票名在D列
        if not stock_name:
            continue

        closing_price = sheet.cell(row=i, column=4).value
        low_price = sheet.cell(row=i + 4, column=4).value
        high_price = sheet.cell(row=i + 5, column=4).value
        change = sheet.cell(row=i + 7, column=4).value

        if closing_price and low_price and high_price:
            stock_data[stock_name] = {
                'closing_price': round(float(closing_price), 2),
                'low_price': round(float(low_price), 2),
                'high_price': round(float(high_price), 2),
                'change': change
            }

    return stock_data

def get_stock_data_today():
    '''
    需要的参数：regularMarketPrice, regularMarketDayHigh, regularMarketDayLow， regularMarketChangePercent 只要raw格式的数值，regularMarketChangePercent改成百分比格式，保留小数点后两位例如 13.51%
    {
    "chart": {
        "result": [
            {
                "meta": {
                    "currency": "HKD",
                    "symbol": "0020.HK",
                    ...
                    "regularMarketPrice": 1.61,
                    "fiftyTwoWeekHigh": 2.35,
                    "fiftyTwoWeekLow": 0.58,
                    "regularMarketDayHigh": 1.66,
                    "regularMarketDayLow": 1.59,
                    "regularMarketVolume": 257744223,
                    "longName": "SENSETIME-W",
                    "shortName": "SENSETIME-W",
                    "chartPreviousClose": 1.63,
                    "previousClose": 1.63,
                    "scale": 3,
                    "priceHint": 3,
    '''
    stock_data = {}

    for code in request_stock_code:
        if not code:
            continue
        encoded_code = quote(code)
        url = f"{request_url}{encoded_code}"
        try:
            response = requests.get(url, headers=headers)
            if response.status_code != 200:
                if response.status_code == 429:
                    retryAfter = response.headers.get('Retry-After', 60)
                    print(f"请求股票{code}失败，状态码：{response.status_code}，重试时间：{retryAfter}秒")
                    time.sleep(int(retryAfter))
                    response = requests.get(url, headers=headers)
                    if response.status_code != 200:
                        print(f"请求股票{code}失败，状态码：{response.status_code}")
                        continue
                else:
                    print(f"请求股票{code}失败，状态码：{response.status_code}")
                    continue
            data = response.json()
            try:
                result = data['chart']['result'][0]
                meta = result['meta']
                stock_code = meta.get('symbol', code)
                regularMarketPrice = meta['regularMarketPrice']
                regularMarketDayHigh = meta['regularMarketDayHigh']
                regularMarketDayLow = meta['regularMarketDayLow']
                previousClose = meta['previousClose']
                stock_data[stock_code] = {
                    'closing_price': round(regularMarketPrice, 3),
                    'low_price': round(regularMarketDayLow, 3),
                    'high_price': round(regularMarketDayHigh, 3),
                    'previous_close': round(previousClose, 3)
                }
                print(stock_data[stock_code])
            except (KeyError, IndexError):
                print(f"请求股票{code}失败，数据格式错误")
                continue
        except requests.exceptions.RequestException as e:
            print(f"请求股票{code}时发生错误: {e}")
            continue

    return stock_data

def update_excel(filename, stock_data):
    '''
    开头为：
    1. 序号：= 行号 - 1
    2. 日期：= 今天日期 格式 `YYYY/MM/DD`
    3. 星期：= 今天星期几，如果星期五，则写5
    EXCEL 每个股票一共8列(无需名字，由request_stock_code决定顺序)
    1. 收盘价：表头表示为股票代码，对应 regularMarketPrice
    2. 日k：无需填写
    3. 周k：无需填写
    4. 月k：无需填写
    5. 低点：对应 regularMarketDayLow
    6. 高点：对应 regularMarketDayHigh
    7. 交易数量：无需填写 (可选列)
    8. 涨幅：对应 (previousClose - regularMarketPrice) / previousClose * 100
    9. 备注：无需填写 (可选列)
    '''
    try:
        wb = load_workbook(filename)
        sheet = wb.active
        stock_col_start = {}
        header_row = sheet[1]
        col = 1
        while col <= sheet.max_column:
            cell_value = header_row[col-1].value
            if cell_value and any(cell_value.startswith(code + " 收盘价") for code in request_stock_code):
                code = next((code for code in request_stock_code if cell_value.startswith(code + " 收盘价")), None)
                if code:
                    stock_col_start[code] = col
            col += 1
    except (FileNotFoundError, InvalidFileException):
        wb = Workbook()
        sheet = wb.active
        headers = ["序号", "日期", "星期"]
        stock_col_start = {}
        current_col = 4
        for idx, code in enumerate(request_stock_code):
            stock_col_start[code] = current_col
            base_columns = [
                f"{code} 收盘价",
                f"{code} 日k",
                f"{code} 周k",
                f"{code} 月k",
                f"{code} 低点",
                f"{code} 高点",
                f"{code} 交易数量",
                f"{code} 涨幅"
            ]
            if "交易数量" in config_list.get(code, []):
                base_columns.remove(f"{code} 交易数量")
            headers.extend(base_columns)
            optional_cols = [col for col in config_list.get(code, []) if col != "交易数量"]
            prefixed_optional_cols = [f"{code} {col}" for col in optional_cols]
            headers.extend(prefixed_optional_cols)
            current_col += len(base_columns) + len(prefixed_optional_cols)
            if idx != len(request_stock_code) - 1:
                headers.append("")
                current_col += 1
        sheet.append(headers)

    today = date.today()
    weekday_num = today.weekday() + 1
    weekday = 5 if weekday_num == 5 else weekday_num
    last_row = sheet.max_row + 1
    sheet.cell(row=last_row, column=1, value=last_row - 1)
    sheet.cell(row=last_row, column=2, value=today.strftime("%Y/%m/%d"))
    sheet.cell(row=last_row, column=3, value=weekday)

    for stock_name, data in stock_data.items():
        try:
            index = request_stock_code.index(stock_name)
            code = request_stock_code[index]
        except ValueError:
            print(f"股票代码 {stock_name} 不在 request_stock_code 列表中，跳过。")
            continue
        base_col = stock_col_start.get(code)
        if not base_col:
            print(f"无法找到股票 {code} 的起始列，跳过。")
            continue
        sheet.cell(row=last_row, column=base_col, value=data['closing_price'])
        sheet.cell(row=last_row, column=base_col + 1, value="")
        sheet.cell(row=last_row, column=base_col + 2, value="")
        sheet.cell(row=last_row, column=base_col + 3, value="")
        sheet.cell(row=last_row, column=base_col + 4, value=data['low_price'])
        sheet.cell(row=last_row, column=base_col + 5, value=data['high_price'])
        if "交易数量" not in config_list.get(code, []):
            change_col = base_col + 7
        else:
            change_col = base_col + 6
        change = ""
        if data.get('previous_close'):
            change_value = (data['closing_price'] - data['previous_close']) / data['previous_close'] * 100
            change = f"{change_value:.3f}%"
        sheet.cell(row=last_row, column=change_col, value=change)
        optional_cols = config_list.get(code, [])
        for i, col in enumerate(optional_cols):
            sheet.cell(row=last_row, column=change_col + 1 + i, value=data.get(col, ''))
    for cell in sheet[last_row]:
        cell.alignment = Alignment(horizontal='center')
    wb.save(filename)
    print(f"数据已成功更新到 {filename}")

def main():
    filename = os.getenv('STOCK_FILENAME', 'stocks.xlsx')
    try:
        stock_data = get_stock_data_today()
        update_excel(filename, stock_data)
    except Exception as e:
        print(f"更新股票数据时发生错误: {e}")

@app.route('/')
def home():
    html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>股票数据管理</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                text-align: center;
                margin-top: 100px;
                background-color: #f4f4f4;
            }
            .button {
                display: inline-block;
                padding: 15px 25px;
                font-size: 16px;
                cursor: pointer;
                text-align: center;
                text-decoration: none;
                outline: none;
                color: #fff;
                background-color: #4CAF50;
                border: none;
                border-radius: 15px;
                box-shadow: 0 9px #999;
                margin: 20px;
            }
            .button:hover {background-color: #45a049}
            .button:active {
                background-color: #45a049;
                box-shadow: 0 5px #666;
                transform: translateY(4px);
            }
            .button-download {
                background-color: #008CBA;
            }
            .button-download:hover {
                background-color: #007bb5;
            }
        </style>
    </head>
    <body>
        <h1>股票数据管理</h1>
        <a href="{{ url_for('update') }}" class="button">更新当天数据</a>
        <a href="{{ url_for('download_file') }}" class="button button-download">获取文件</a>
        <a href="{{ url_for('logout') }}" class="button" style="background-color: #f44336;">注销</a>
    </body>
    </html>
    '''
    return render_template_string(html)

@app.route('/update', methods=['GET'])
def update():
    filename = os.getenv('STOCK_FILENAME', 'stocks.xlsx')
    # 如果本地文件存在，则先删除
    if os.path.exists(filename):
        os.remove(filename)
    try:
        stock_data = get_stock_data_today()
        update_excel(filename, stock_data)
        return jsonify({"status": "success", "message": f"数据已更新到 {filename}"}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": f"更新数据时发生错误: {e}"}), 500

@app.route('/download', methods=['GET'])
def download_file():
    filename = os.getenv('STOCK_FILENAME', 'stocks.xlsx')
    directory = os.path.abspath('.')
    try:
        return send_from_directory(directory, filename, as_attachment=True)
    except Exception as e:
        return jsonify({"status": "error", "message": f"下载文件时发生错误: {e}"}), 500

# 添加错误处理装饰器
@app.errorhandler(503)
def service_unavailable(e):
    return jsonify(error="服务暂时不可用，请稍后再试"), 503

# 修改主运行代码，添加更多的错误处理和日志
if __name__ == "__main__":
    shanghai = pytz.timezone('Asia/Shanghai')
    scheduler = BackgroundScheduler(timezone=shanghai)
    scheduler.add_job(update, 'cron', hour=0, minute=0)
    
    try:
        scheduler.start()
        logger.info("应用程序启动中...")
        # 关闭调试模式，使用生产配置
        app.run(host='0.0.0.0', port=5000, debug=False)
    except Exception as e:
        logger.error(f"启动服务器时发生错误: {e}")
    finally:
        scheduler.shutdown()
